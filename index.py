from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import requests
import csv
from bs4 import BeautifulSoup as bs
import io
import urllib3
import os

        
class Book:
    def __init__(self, title, author, ratingVal, genre, pages, nrating, nreview, imgCover):
        self.title = title
        self.author = author
        self.ratingVal = ratingVal
        self.genre = genre,
        self.nofPages = pages
        self.nrating = nrating
        self.nreview = nreview
        self.imgCover = imgCover

def scrape_and_run(mybook):
    # scrape on goodreads.com using desire genre type or key word
    # and save the titles and autors in a csv file

    page = requests.get("https://www.goodreads.com/search?q=" + mybook)
    soup = bs(page.content, 'html.parser')
    titles = soup.find('a', class_='bookTitle')
    authors = soup.find('a', class_='authorName')

    image_dir = os.getcwd() + "/images/mybook"

    ## check if the desire genre path exists
    ## create a new one if it doesnt
    if not os.path.exists(image_dir):
        os.makedirs(image_dir)
    try:
        book_page = requests.get("https://www.goodreads.com" + titles['href'])
        
        soup = bs(book_page.content, 'html.parser')
        ratingVal = soup.find('span', itemprop='ratingValue').get_text()
        genres = soup.find_all('a', class_='actionLinkLite bookPageGenreLink')
        genre = ""

        for g in genres:
            genre+= (g.get_text())
        
        http = urllib3.PoolManager()
        imgUrl = soup.find('img', id='coverImage')
        r = http.request('GET', imgUrl['src'])
        image_file = io.BytesIO(r.data)
        image_file.width = 4
        image_file.height = 5
        img = Image(image_file)
        
        try:
            nofPages = soup.find('span', itemprop='numberOfPages').get_text()
        except AttributeError as exc:
            print(exc)
            return None    
        nrating = soup.find('meta', itemprop='ratingCount')['content']
        nreview = soup.find('meta', itemprop='reviewCount')['content']
        
        title_name = titles.get_text()
        author_name = authors.get_text()
        # print(nofPages)
        # print(nrating)
        # print(nreview)
        book = Book(title_name, author_name, ratingVal, genre, nofPages, nrating, nreview, img)
        return book
    except TypeError as exc:
        print(exc)
        return None    

if __name__ == '__main__':
    filename = 'books2.xlsx'
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    for row in range(1, 20):
        title = sheet.cell(column=1, row=row).value
        if title != None:
            book = title.lower() # input case lowered
            print(book)
            resBook = scrape_and_run(book)
            if resBook != None:
                sheet.cell(column=2, row=row).value = resBook.author
                sheet.cell(column=3, row=row).value = resBook.ratingVal
                sheet.cell(column=4, row=row).value = resBook.genre[0]
                sheet.cell(column=5, row=row).value = resBook.nofPages
                sheet.cell(column=6, row=row).value = resBook.nrating
                sheet.cell(column=7, row=row).value = resBook.nreview

                loc ='H'+ str(row)
                print(loc)
                resBook.imgCover.anchor = loc
                sheet.add_image(resBook.imgCover)
    workbook.save(filename=filename)
    # scrape_and_run("chiến binh cầu vồng")
            

